import openpyxl
import requests, json
import xlsxwriter
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

import os
li = []
wrkbk = openpyxl.load_workbook("C:/Users/masrafe/Desktop/selenium/Excel.xlsx")
sh = wrkbk.active
for row in sh.iter_rows(min_row=3, min_col=3, max_row=10, max_col=3):
	for cell in row:
		li.append(cell.value)

l = []
s = []     

for i in li:
    headers = {
    "User-Agent":
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.102 Safari/537.36 Edge/18.19582"}
    a = 'http://google.com/complete/search?client=chrome&q='+ i
    response = requests.get(a, headers=headers)
    result = json.loads(response.text)[1]
    long = max(result, key=len)
    short = min(result,key=len)
    s.append(short)
    l.append(long)

dt = datetime.now()
day =dt.strftime('%A')
date_time = datetime.now()

current_row = 2
column_number = 4
for i in l:
    current_row += 1
    sh.cell(row=current_row, column=column_number).value = i
    os.chmod('C:/Users/masrafe/Desktop/selenium/Excel.xlsx', 0o777)

    wrkbk.save('C:/Users/masrafe/Desktop/selenium/Excel.xlsx')
current_row = 2
column_number = 5
for i in s:
    current_row += 1
    sh.cell(row=current_row, column=column_number).value = i
    sh.cell(row=current_row, column=column_number+1).value = day
    sh.cell(row=current_row, column=column_number+2).value = date_time
    os.chmod('C:/Users/masrafe/Desktop/selenium/Excel.xlsx', 0o777)

    wrkbk.save('C:/Users/masrafe/Desktop/selenium/Excel.xlsx')

    








	

